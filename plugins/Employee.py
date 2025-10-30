from __future__ import annotations
from typing import Any, Dict, List, Tuple
from pathlib import Path
import json
import sys

try:
    from openpyxl import load_workbook
except ImportError:
    sys.stderr.write("Error: openpyxl is required. Install it with: pip install openpyxl\n")
    raise

EXCEL_ID_COL = "מס'"
EXCEL_MANAGER_COL = "שם פרטי"
EXCEL_PHONE_COL = "טלפון נייד"

def _safe_int(x: Any) -> int | None:
    try:
        if x is None or x == "":
            return None
        return int(float(str(x).strip()))
    except Exception:
        return None

def _normalize_phone(s: Any) -> str:
    if s is None:
        return ""
    return str(s).strip()

def _prepare_headers(values_first_row: List[Any]) -> Tuple[List[str], List[int]]:
    counts: Dict[str, int] = {}
    headers: List[str] = []
    keep_indices: List[int] = []
    for idx, h in enumerate(values_first_row or []):
        name = str(h).strip() if h is not None else ""
        if not name:
            continue
        base = name
        n = counts.get(base, 0)
        if n:
            name = f"{base}_{n+1}"
        counts[base] = n + 1
        headers.append(name)
        keep_indices.append(idx)
    return headers, keep_indices

def _excel_to_filtered_rows(xlsx_path: Path, sheet_name: str | None) -> List[Dict[str, Any]]:
    wb = load_workbook(filename=str(xlsx_path), data_only=True, read_only=True)
    ws = wb[sheet_name] if sheet_name else wb.worksheets[0]
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return []
    headers, keep_indices = _prepare_headers(header_row or [])
    cols = {h: i for i, h in enumerate(headers)}
    need = [EXCEL_ID_COL, EXCEL_MANAGER_COL, EXCEL_PHONE_COL]
    for k in need:
        if k not in cols:
            cols[k] = None
    out: List[Dict[str, Any]] = []
    for row in rows_iter:
        if row is None:
            continue
        item: Dict[str, Any] = {}
        for k in need:
            ci = cols[k]
            val = row[ci] if (ci is not None and ci < len(row)) else None
            if k == EXCEL_ID_COL:
                item[k] = _safe_int(val)
            elif k == EXCEL_PHONE_COL:
                item[k] = _normalize_phone(val)
            else:
                item[k] = "" if val is None else str(val).strip()
        if item.get(EXCEL_ID_COL) is not None:
            out.append(item)
    return out

def prepare() -> Dict[str, Any]:
    base = Path(__file__).resolve().parent.parent
    xlsx_path = base / "data" / "EmployeesData.xlsx"
    rows = _excel_to_filtered_rows(xlsx_path, None)
    return {
        "employees_json_path": str(base / "employeesData.json"),
        "data": rows,
        "cleanup": True
    }

def _get_branches_ref(printers: Dict[str, Any]) -> tuple[str, List[Dict[str, Any]] | None]:
    for k, v in printers.items():
        if isinstance(v, list) and k.lower() == "branches_grouped":
            return k, v
    return "", None

def _load_json(path: str | Path) -> Any:
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)

def run(printers: Dict[str, Any], employees_json_path: str | None = None) -> Dict[str, Any]:
    if employees_json_path is None:
        base = Path(__file__).resolve().parent.parent
        employees_json_path = str(base / "employeesData.json")
    employees = _load_json(employees_json_path)
    key, branches = _get_branches_ref(printers)
    if not branches or not isinstance(employees, list):
        return printers

    lut: Dict[int, Tuple[str, str]] = {}
    for r in employees:
        if not isinstance(r, dict):
            continue
        branch_id = _safe_int(r.get(EXCEL_ID_COL))
        if branch_id is None:
            continue
        m = "" if r.get(EXCEL_MANAGER_COL) is None else str(r.get(EXCEL_MANAGER_COL)).strip()
        p = _normalize_phone(r.get(EXCEL_PHONE_COL))
        prev = lut.get(branch_id)
        if prev:
            m = m or prev[0]
            p = p or prev[1]
        lut[branch_id] = (m, p)

    if not lut:
        return printers

    for i, obj in enumerate(branches):
        if not isinstance(obj, dict):
            continue
        branch_id = _safe_int(obj.get("ID") or obj.get("Id") or obj.get("id"))
        if branch_id is None:
            continue
        if branch_id in lut:
            manager, phone = lut[branch_id]
            existing = obj.get("storeInfo")
            if not isinstance(existing, dict):
                existing = {}
            new_fields: Dict[str, Any] = {}
            if manager and (existing.get("Manager") in (None, "")):
                new_fields["Manager"] = manager
            if phone and (existing.get("Phone") in (None, "")):
                new_fields["Phone"] = phone
            if new_fields:
                si = dict(existing)
                for k_add, v_add in new_fields.items():
                    si[k_add] = v_add
                obj["storeInfo"] = si
            branches[i] = obj

    printers[key] = branches
    return printers
