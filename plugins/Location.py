from __future__ import annotations
from typing import Any, Dict, List, Tuple
from pathlib import Path
import json
import sys
import re

try:
    from openpyxl import load_workbook
except ImportError:
    sys.stderr.write("Error: openpyxl is required. Install it with: pip install openpyxl\n")
    raise

EXCEL_PATH = r"\\st-filea\St-SystemIT\IT\Stores\בזק\stores\קווים בחנויות.xlsx"
EXCEL_BRANCH_ID_COL = "מס' סניף"
EXCEL_ADDRESS_COL = "כתובת"
EXCEL_PRIMARY_DESC_COL = "תאור שרות ראשי"
EXCEL_SECONDARY_DESC_COL = "תאור שרות משני"
EXCEL_SUBSCR_NUM_COL = "מספר מנוי"

def _safe_int(x: Any) -> int | None:
    try:
        if x is None or x == "":
            return None
        return int(float(str(x).strip()))
    except Exception:
        return None

def _prepare_headers(first_row: List[Any]) -> tuple[List[str], List[int]]:
    counts: Dict[str, int] = {}
    headers: List[str] = []
    keep: List[int] = []
    for idx, h in enumerate(first_row or []):
        name = str(h).strip() if h is not None else ""
        if not name:
            continue
        base = name
        n = counts.get(base, 0)
        if n:
            name = f"{base}_{n+1}"
        counts[base] = n + 1
        headers.append(name)
        keep.append(idx)
    return headers, keep

def _row_is_empty(row: tuple[Any, ...] | None, indices: List[int]) -> bool:
    if row is None:
        return True
    for i in indices:
        v = row[i] if i < len(row) else None
        if v is None:
            continue
        if not (isinstance(v, str) and v.strip() == ""):
            return False
    return True

def _excel_to_rows(path: str, sheet_name: str | None = None) -> List[Dict[str, Any]]:
    wb = load_workbook(filename=path, data_only=True, read_only=True)
    ws = wb[sheet_name] if sheet_name else wb.worksheets[0]
    it = ws.iter_rows(values_only=True)
    try:
        header_row = next(it)
    except StopIteration:
        return []
    headers, keep = _prepare_headers(header_row or [])
    if not headers:
        return []
    out: List[Dict[str, Any]] = []
    for row in it:
        if _row_is_empty(row, keep):
            continue
        item: Dict[str, Any] = {}
        for col_idx, h in zip(keep, headers):
            val = row[col_idx] if row is not None and col_idx < len(row) else None
            item[h] = val
        out.append(item)
    return out

def _get_branches_ref(printers: Dict[str, Any]) -> tuple[str, List[Dict[str, Any]] | None]:
    for k, v in printers.items():
        if isinstance(v, list) and k.lower() == "branches_grouped":
            return k, v
    return "", None

def _load_json(path: str | Path) -> Any:
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)

def _split_postal(addr: str) -> Tuple[str, str | None]:
    if not addr:
        return "", None
    m = re.search(r'(\d{7})\s*$', addr)
    if not m:
        return addr.strip(), None
    postal = m.group(1)
    cleaned = re.sub(r'[\s,;:-]*\d{7}\s*$', "", addr).rstrip(" ,;:-")
    return cleaned.strip(), postal

def _norm_text(x: Any) -> str | None:
    if x is None:
        return None
    s = str(x).strip()
    if not s:
        return None
    s = re.sub(r"\s+", " ", s)
    return s

def prepare() -> Dict[str, Any]:
    base = Path(__file__).resolve().parent.parent
    rows = _excel_to_rows(EXCEL_PATH, None)
    return {
        "employees_json_path": str(base / "employeesData_location.json"),
        "data": rows,
        "cleanup": True
    }

def run(printers: Dict[str, Any], employees_json_path: str | None = None) -> Dict[str, Any]:
    if employees_json_path is None:
        base = Path(__file__).resolve().parent.parent
        employees_json_path = str(base / "employeesData_location.json")
    employees = _load_json(employees_json_path)
    key, branches = _get_branches_ref(printers)
    if not branches or not isinstance(employees, list):
        return printers

    addr_by_branch: Dict[int, str] = {}
    for r in employees:
        if not isinstance(r, dict):
            continue
        bid = _safe_int(r.get(EXCEL_BRANCH_ID_COL))
        if bid is None:
            continue
        addr = "" if r.get(EXCEL_ADDRESS_COL) is None else str(r.get(EXCEL_ADDRESS_COL)).strip()
        if addr:
            addr_by_branch[bid] = addr

    pair_order_by_branch: Dict[int, List[Tuple[str, str]]] = {}
    subs_by_branch: Dict[int, Dict[Tuple[str, str], List[str]]] = {}
    seen_pairs_by_branch: Dict[int, set[Tuple[str, str]]] = {}

    for r in employees:
        if not isinstance(r, dict):
            continue
        bid = _safe_int(r.get(EXCEL_BRANCH_ID_COL))
        if bid is None:
            continue

        p = _norm_text(r.get(EXCEL_PRIMARY_DESC_COL))
        s = _norm_text(r.get(EXCEL_SECONDARY_DESC_COL))
        if p is None and s is None:
            continue
        pair = (p or "", s or "")

        subs_map = subs_by_branch.setdefault(bid, {})
        subs_list = subs_map.setdefault(pair, [])
        sub_num = _norm_text(r.get(EXCEL_SUBSCR_NUM_COL))
        if sub_num is not None:
            subs_list.append(sub_num)

        seen = seen_pairs_by_branch.setdefault(bid, set())
        if pair not in seen:
            seen.add(pair)
            pair_order_by_branch.setdefault(bid, []).append(pair)

    for i, obj in enumerate(branches):
        if not isinstance(obj, dict):
            continue
        branch_id = _safe_int(obj.get("ID") or obj.get("Id") or obj.get("id"))
        if branch_id is None:
            continue

        si = obj.get("storeInfo")
        if not isinstance(si, dict):
            si = {}

        addr = addr_by_branch.get(branch_id, "")
        if addr:
            location, postal = _split_postal(addr)
            si["Location"] = location
            si["Postal"] = postal

        order = pair_order_by_branch.get(branch_id, [])
        subs_map = subs_by_branch.get(branch_id, {})

        def _make_desc(pair: Tuple[str, str]) -> Dict[str, str]:
            p, s = pair
            nums = subs_map.get(pair, [])
            line_id = nums[0] if nums else ""
            return {
                "LineID": str(line_id),
                "PrimaryDescription": p,
                "SecondayDescription": s
            }

        if len(order) >= 1:
            si["firstDescription"] = _make_desc(order[0])
        if len(order) >= 2:
            si["secondDescription"] = _make_desc(order[1])

        if "firstDisc" in si:
            del si["firstDisc"]
        if "secondDisc" in si:
            del si["secondDisc"]

        obj["storeInfo"] = si
        branches[i] = obj

    printers[key] = branches
    return printers
