import argparse, json, math, shutil
from pathlib import Path
from datetime import date, datetime

try:
    import pandas as pd
    import numpy as np
    _HAS_PANDAS = True
except Exception:
    _HAS_PANDAS = False


SHEET_SPECS = {
    "Company_Grouped": {
        "keys": ["ID", "Floor", "Printer IP", "Type", "Serial", "Comment"],
        "cols": 6,
    },
    "Branches_Grouped": {
        "keys": ["ID", "Name", "Printer IP", "BO IP", "Type", "Serial", "Comment"],
        "cols": 7,
    },
}

def _norm(s):
    if s is None:
        return ""
    return " ".join(str(s).strip().lower().split())

def json_serializer(obj):
    try:
        import pandas as _pd
    except Exception:
        _pd = None
    if (_pd is not None and isinstance(obj, getattr(_pd, "Timestamp", ()))) or isinstance(obj, (datetime, date)):
        return obj.isoformat()
    try:
        import numpy as _np
        if isinstance(obj, _np.integer):
            return int(obj)
        if isinstance(obj, _np.floating):
            val = float(obj)
            return None if math.isnan(val) else val
        if isinstance(obj, _np.bool_):
            return bool(obj)
    except Exception:
        pass
    return str(obj)

def _build_header_map(headers, limit):
    normed = [_norm(h) for h in headers[:limit]]
    return {i: n for i, n in enumerate(normed)}

def _choose_indices_for_sheet(sheet_name, header_map):
    spec = SHEET_SPECS[sheet_name]
    keys = spec["keys"]
    limit = spec["cols"]
    idx_for_key = [-1] * len(keys)
    taken = set()
    for k_i, key in enumerate(keys):
        target = _norm(key)
        found = -1
        for col_i in range(limit):
            if col_i in header_map and header_map[col_i] == target and col_i not in taken:
                found = col_i
                break
        idx_for_key[k_i] = found
        if found >= 0:
            taken.add(found)
    return idx_for_key

def _row_to_record(row_vals, keys, mapping, limit):
    vals = [(row_vals[i] if i < len(row_vals) else None) for i in range(limit)]
    out = {}
    for i, k in enumerate(keys):
        src = mapping[i] if mapping[i] >= 0 else i
        v = vals[src] if src < len(vals) else None
        if isinstance(v, float) and math.isnan(v):
            v = None
        out[k] = v
    if all(v in (None, "", []) for v in out.values()):
        return None
    return out

def load_with_pandas_hybrid(xl_path: Path, sheets):
    import pandas as pd
    import numpy as np
    out = {}
    for sheet in sheets:
        spec = SHEET_SPECS[sheet]
        keys, limit = spec["keys"], spec["cols"]
        df = pd.read_excel(xl_path, sheet_name=sheet, header=0, engine="openpyxl")
        headers = list(df.columns)
        header_map = _build_header_map(headers, limit)
        mapping = _choose_indices_for_sheet(sheet, header_map)
        df = df.iloc[:, :max(limit, len(headers))].replace({np.nan: None})
        records = []
        for _, row in df.iterrows():
            rec = _row_to_record(list(row.values), keys, mapping, limit)
            if rec is not None:
                records.append(rec)
        out[sheet] = records
    return out

def load_with_openpyxl_hybrid(xl_path: Path, sheets):
    from openpyxl import load_workbook
    wb = load_workbook(filename=str(xl_path), data_only=True, read_only=False)
    out = {}
    for sheet in sheets:
        spec = SHEET_SPECS[sheet]
        keys, limit = spec["keys"], spec["cols"]
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            out[sheet] = []
            continue
        headers = list(rows[0]) if rows else []
        header_map = _build_header_map(headers, limit)
        mapping = _choose_indices_for_sheet(sheet, header_map)
        records = []
        for row in rows[1:]:
            row_vals = list(row) if row else []
            rec = _row_to_record(row_vals, keys, mapping, limit)
            if rec is not None:
                records.append(rec)
        out[sheet] = records
    return out

def _resolve_xlsm(p: str | Path) -> Path:
    p = Path(p).expanduser().resolve()
    if p.suffix == "":
        p = p.with_suffix(".xlsm")
    if p.suffix.lower() != ".xlsm":
        raise SystemExit(f"Only .xlsm files are supported (got {p.suffix}).")
    return p

def copy_draft_to_prod(draft_path: Path, prod_path: Path):
    prod_path.parent.mkdir(parents=True, exist_ok=True)
    if draft_path.samefile(prod_path):
        # Nothing to do; already the same file
        return
    shutil.copy2(draft_path, prod_path)

def main():
    parser = argparse.ArgumentParser()
    # New: explicit draft/prod paths. Defaults implement your requested behavior.
    parser.add_argument("--draft", default="data/printersDraft.xlsm", help="Path to the draft xlsm")
    parser.add_argument("--prod", default="printers.xlsm", help="Path to the production xlsm to overwrite")
    parser.add_argument("--output","-o", default=None, help="JSON output path; defaults to <prod>.json")
    parser.add_argument("--sheets", nargs="+", default=["Company_Grouped","Branches_Grouped"])
    args = parser.parse_args()

    draft_path = _resolve_xlsm(args.draft)
    if not draft_path.exists():
        raise SystemExit(f"Draft Excel not found (expected .xlsm): {draft_path}")

    prod_path = _resolve_xlsm(args.prod)

    # 1) Overwrite prod with the draft
    copy_draft_to_prod(draft_path, prod_path)

    # 2) Convert (post-overwrite) to JSON, defaulting next to prod
    out_path = Path(args.output).expanduser().resolve() if args.output else prod_path.with_suffix(".json")

    for s in args.sheets:
        if s not in SHEET_SPECS:
            raise SystemExit(f"Unsupported sheet: {s}")

    # Read from the production file (now updated from the draft)
    if _HAS_PANDAS:
        data = load_with_pandas_hybrid(prod_path, args.sheets)
    else:
        data = load_with_openpyxl_hybrid(prod_path, args.sheets)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2, default=json_serializer)

    print(f"Overwrote {prod_path} from {draft_path}")
    print(f"Wrote {out_path}")

if __name__ == "__main__":
    main()
