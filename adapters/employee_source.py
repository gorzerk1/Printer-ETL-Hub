from __future__ import annotations
from pathlib import Path
from typing import Any, Dict, List
from openpyxl import load_workbook

def read_employees_xlsx(path: str | Path) -> List[Dict[str, Any]]:
    p = Path(path)
    if not p.exists():
        return []
    wb = load_workbook(filename=str(p), data_only=True, read_only=True)
    ws = wb.worksheets[0]
    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        return []
    headers = [str(h or "").strip().lower() for h in header]
    def idx(opts: List[str]) -> int | None:
        for n in opts:
            low = n.lower()
            if low in headers:
                return headers.index(low)
        return None
    id_i = idx(["id","branch id","מספר סניף","מס'"])
    name_i = idx(["name","contact","contacts name","manager","שם איש קשר","שם פרטי"])
    phone_i = idx(["phone","contacts phone","טלפון","טלפון נייד"])
    items: List[Dict[str, Any]] = []
    for row in rows:
        vals = list(row) if row else []
        def at(i):
            return None if i is None or i >= len(vals) else vals[i]
        rid = at(id_i)
        nm = at(name_i)
        ph = at(phone_i)
        if rid is None and nm is None and ph is None:
            continue
        items.append({
            "id": "" if rid is None else str(rid).strip(),
            "name": "" if nm is None else str(nm).strip(),
            "phone": "" if ph is None else str(ph).strip(),
        })
    return items
