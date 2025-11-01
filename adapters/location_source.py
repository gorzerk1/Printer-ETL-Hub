from __future__ import annotations
from pathlib import Path
from typing import Any, Dict, List
from openpyxl import load_workbook

def _prepare_headers(first_row: List[Any]) -> tuple[List[str], List[int]]:
    counts: Dict[str, int] = {}
    headers: List[str] = []
    keep: List[int] = []
    for idx, h in enumerate(first_row or []):
        name = str(h).strip() if h is not None else ""
        if not name:
            continue
        base = name
        n = counts.get(base, 0)
        if n:
            name = f"{base}_{n+1}"
        counts[base] = n + 1
        headers.append(name)
        keep.append(idx)
    return headers, keep

def _row_is_empty(row: tuple[Any, ...] | None, indices: List[int]) -> bool:
    if row is None:
        return True
    for i in indices:
        v = row[i] if i < len(row) else None
        if v is None:
            continue
        if not (isinstance(v, str) and v.strip() == ""):
            return False
    return True

def read_locations_xlsx(path: str | Path, sheet: str | None = None) -> List[Dict[str, Any]]:
    p = Path(path)
    if not p.exists():
        return []
    wb = load_workbook(filename=str(p), data_only=True, read_only=True)
    ws = wb[sheet] if sheet else wb.worksheets[0]
    it = ws.iter_rows(values_only=True)
    try:
        header_row = next(it)
    except StopIteration:
        return []
    headers, keep = _prepare_headers(list(header_row or []))
    if not headers:
        return []
    out: List[Dict[str, Any]] = []
    for row in it:
        if _row_is_empty(row, keep):
            continue
        item: Dict[str, Any] = {}
        for col_idx, h in zip(keep, headers):
            val = row[col_idx] if row is not None and col_idx < len(row) else None
            item[h] = val
        out.append(item)
    return out
