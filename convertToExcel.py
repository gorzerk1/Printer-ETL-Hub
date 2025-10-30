from __future__ import annotations
import sys, json, re, os, subprocess
from pathlib import Path
from typing import Dict, Any, Optional, Tuple, List, Set
from datetime import datetime
import openpyxl

_ILLEGAL_XML_CHARS_RE = re.compile(r'[\x00-\x08\x0B\x0C\x0E-\x1F]')

def sanitize_excel_value(val):
    if val is None:
        return None
    if isinstance(val, str):
        return _ILLEGAL_XML_CHARS_RE.sub("", val)
    return val

def canonicalize_id(v) -> Optional[str]:
    if v is None:
        return None
    try:
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            return str(int(v))
    except Exception:
        pass
    s = str(v).strip().replace("\n", " ").replace("\r", " ").strip()
    return s

def normalize_color(name: str) -> Optional[str]:
    if not name:
        return None
    n = str(name).strip().lower()
    if "black" in n or n == "k":
        return "Black"
    if "cyan" in n or n == "c":
        return "Cyan"
    if "magenta" in n or n == "m":
        return "Magenta"
    if "yellow" in n or n == "y":
        return "Yellow"
    return None

def _status_online_offline(raw) -> str:
    if raw is None:
        return "offline"
    s = str(raw).strip().lower()
    if not s:
        return "offline"
    online_keys = ("online","ready","idle","sleep","printing","working","active","ok","connected")
    offline_keys = ("offline","down","disconnected","error","unknown","not reachable","unreachable","no connection","disabled")
    if any(k in s for k in online_keys):
        return "online"
    if any(k in s for k in offline_keys):
        return "offline"
    if "off" in s:
        return "offline"
    if "on" in s:
        return "online"
    return "offline"

def dash_if_blank(val):
    if val is None:
        return "-"
    if isinstance(val, str) and not val.strip():
        return "-"
    return val

def _pick_printer_error(pinfo: Dict[str, Any]) -> Tuple[Optional[str], Optional[str]]:
    err = pinfo.get("printerError")
    if not err:
        return None, None
    candidate = None
    if isinstance(err, dict):
        candidate = err
    elif isinstance(err, list) and err:
        for x in err:
            if isinstance(x, dict) and (x.get("problem") or x.get("severity")):
                candidate = x
                break
    if not isinstance(candidate, dict):
        return None, None
    problem = candidate.get("problem")
    severity = candidate.get("severity")
    problem = None if problem is None else str(problem)
    severity = None if severity is None else str(severity)
    return problem, severity

def _join_toner_types(tt) -> Optional[str]:
    if tt is None:
        return None
    if isinstance(tt, (list, tuple)):
        parts: List[str] = []
        seen: Set[str] = set()
        for x in tt:
            if x is None:
                continue
            s = str(x).strip()
            if not s or s in seen:
                continue
            parts.append(s)
            seen.add(s)
        return ", ".join(parts) if parts else None
    s = str(tt).strip()
    return s or None

def extract_info(prn: Dict[str, Any]) -> Dict[str, Any]:
    info = {"Status": None, "Black": None, "Cyan": None, "Magenta": None, "Yellow": None, "Error": None, "Severity": None, "Toner Type": None}
    pinfo = prn.get("printerInfo") or {}
    status_val = pinfo.get("status")
    if status_val is None:
        carts = pinfo.get("cartridges") or []
        if carts and isinstance(carts, list):
            first = carts[0] or {}
            status_val = first.get("status")
    info["Status"] = _status_online_offline(status_val)
    carts = pinfo.get("cartridges") or []
    for cart in carts:
        try:
            cname = normalize_color(cart.get("cartridge"))
            if not cname:
                continue
            rp = cart.get("remaining_percent")
            if rp is None:
                value = None
            else:
                try:
                    value = float(rp)
                    if hasattr(value, "is_integer") and value.is_integer():
                        value = int(value)
                except Exception:
                    value = rp
            info[cname] = value if info.get(cname) in (None, "-") else info.get(cname)
        except Exception:
            continue
    problem, severity = _pick_printer_error(pinfo)
    info["Error"] = problem
    info["Severity"] = severity
    toner_types = _join_toner_types(pinfo.get("tonerType"))
    info["Toner Type"] = toner_types
    return info

def _iter_printers(obj):
    if isinstance(obj, dict):
        if "ID" in obj and isinstance(obj.get("printerInfo"), dict):
            yield obj
        for v in obj.values():
            yield from _iter_printers(v)
    elif isinstance(obj, list):
        for x in obj:
            yield from _iter_printers(x)

def _score_info(info: Dict[str, Any]) -> int:
    keys = ("Toner Type","Black","Cyan","Magenta","Yellow","Status","Error","Severity")
    s = 0
    for k in keys:
        v = info.get(k)
        if v not in (None, "-"):
            s += 1
    return s

def build_id_map(data: Dict[str, Any]) -> Dict[str, Dict[str, Any]]:
    id_map: Dict[str, Dict[str, Any]] = {}
    for obj in _iter_printers(data):
        cid = canonicalize_id(obj.get("ID"))
        if not cid:
            continue
        info = extract_info(obj)
        if cid not in id_map or _score_info(info) > _score_info(id_map[cid]):
            id_map[cid] = info
    return id_map

def find_header_row_and_map(ws) -> Tuple[Optional[int], Dict[str, int]]:
    max_scan_rows = min(max(ws.max_row, 1), 20)
    expected = {"id","status","black","cyan","magenta","yellow","error","severity","toner type","type"}
    best_row = None
    best_score = -1
    best_map: Dict[str, int] = {}
    for r in range(1, max_scan_rows + 1):
        row_map: Dict[str, int] = {}
        score = 0
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v is None:
                continue
            name = str(v).strip()
            if not name:
                continue
            row_map[name] = c
            if name.strip().lower() in expected:
                score += 1
        if "id" in {k.strip().lower() for k in row_map} and score > best_score:
            best_row = r
            best_score = score
            best_map = row_map
    if best_row is None:
        return None, {}
    return best_row, best_map

def ensure_columns(ws, header_row: int, header_map: Dict[str, int], required_cols: List[str]) -> Dict[str, int]:
    last_col = ws.max_column
    lower_map = {k.strip().lower(): v for k, v in header_map.items()}
    if "toner type" not in lower_map:
        type_cols: List[int] = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(header_row, c).value
            if isinstance(v, str) and v.strip().lower() == "type":
                type_cols.append(c)
        if type_cols:
            ws.cell(header_row, type_cols[0], "Toner Type")
            header_map["Toner Type"] = type_cols[0]
            lower_map["toner type"] = type_cols[0]
    for col_name in required_cols:
        if col_name.strip().lower() in lower_map:
            continue
        last_col += 1
        ws.cell(header_row, last_col, col_name)
        header_map[col_name] = last_col
        lower_map[col_name.strip().lower()] = last_col
    return header_map

def update_sheet(ws, id_map: Dict[str, Dict[str, Any]]) -> int:
    header_row, header_map = find_header_row_and_map(ws)
    if not header_row:
        return 0
    required_cols = ["Status", "Black", "Cyan", "Magenta", "Yellow", "Error", "Severity", "Toner Type"]
    header_map = ensure_columns(ws, header_row, header_map, required_cols)
    lower_map = {k.strip().lower(): v for k, v in header_map.items()}
    id_col = lower_map.get("id")
    if not id_col:
        return 0
    updates = 0
    for r in range(header_row + 1, ws.max_row + 1):
        rid_val = ws.cell(r, id_col).value
        cid = canonicalize_id(rid_val)
        if not cid:
            continue
        info = id_map.get(cid)
        if not info:
            continue
        for name in ["Status", "Black", "Cyan", "Magenta", "Yellow", "Error", "Severity", "Toner Type"]:
            cidx = lower_map.get(name.strip().lower())
            if not cidx:
                continue
            ws.cell(r, cidx, sanitize_excel_value(dash_if_blank(info.get(name))))
        updates += 1
    return updates

def _default_employees_json(project_root: Path) -> Path:
    return project_root / "data" / "employeesData.json"

def _iter_dicts_with_key(obj, key: str):
    if isinstance(obj, dict):
        if key in obj:
            yield obj
        for v in obj.values():
            yield from _iter_dicts_with_key(v, key)
    elif isinstance(obj, list):
        for x in obj:
            yield from _iter_dicts_with_key(x, key)

def build_employees_index(employees_json_obj: Any) -> Dict[str, Dict[str, Any]]:
    index: Dict[str, Dict[str, Any]] = {}
    if isinstance(employees_json_obj, list):
        for rec in employees_json_obj:
            if not isinstance(rec, dict):
                continue
            emp_id = canonicalize_id(rec.get("מס'"))
            if not emp_id:
                continue
            name = rec.get("שם פרטי")
            phone = rec.get("טלפון נייד")
            index[emp_id] = {"name": name, "phone": phone}
        if index:
            return index
    for rec in _iter_dicts_with_key(employees_json_obj, "מס'"):
        try:
            emp_id = canonicalize_id(rec.get("מס'"))
            if not emp_id:
                continue
            name = rec.get("שם פרטי")
            phone = rec.get("טלפון נייד")
            index[emp_id] = {"name": name, "phone": phone}
        except Exception:
            continue
    if not index and isinstance(employees_json_obj, dict):
        for k, v in employees_json_obj.items():
            try:
                emp_id = canonicalize_id(k)
                if not emp_id:
                    continue
                if isinstance(v, dict):
                    name = v.get("שם פרטי")
                    phone = v.get("טלפון נייד")
                    if name is not None or phone is not None:
                        index[emp_id] = {"name": name, "phone": phone}
            except Exception:
                continue
    return index

def update_branches_grouped(ws, employees_index: Dict[str, Dict[str, Any]]) -> int:
    if ws.title != "Branches_Grouped":
        return 0
    header_row, header_map = find_header_row_and_map(ws)
    if not header_row:
        return 0
    required_cols = ["Contacts Name", "Contacts Phone"]
    header_map = ensure_columns(ws, header_row, header_map, required_cols)
    lower_map = {k.strip().lower(): v for k, v in header_map.items()}
    id_col = lower_map.get("id")
    if not id_col:
        return 0
    name_col = lower_map.get("contacts name")
    phone_col = lower_map.get("contacts phone")
    updates = 0
    for r in range(header_row + 1, ws.max_row + 1):
        rid_val = ws.cell(r, id_col).value
        cid = canonicalize_id(rid_val)
        if not cid:
            continue
        emp = employees_index.get(cid)
        if not emp:
            continue
        if name_col:
            ws.cell(r, name_col, sanitize_excel_value(dash_if_blank(emp.get("name"))))
        if phone_col:
            ws.cell(r, phone_col, sanitize_excel_value(dash_if_blank(emp.get("phone"))))
        updates += 1
    return updates

def _script_base_dir() -> Path:
    return Path(__file__).resolve().parent

def _logs_dir() -> Path:
    return _script_base_dir() / "logs" / "printerExcel"

def _safe_timestamp_filename(ext: str = ".xlsm") -> str:
    ts = datetime.now().strftime("%Y-%m-%d %H-%M-%S")
    return f"{ts}{ext}"

def _unique_path(base_dir: Path, name: str) -> Path:
    target = base_dir / name
    if not target.exists():
        return target
    stem, suf = os.path.splitext(name)
    i = 1
    while True:
        cand = base_dir / f"{stem}_{i}{suf}"
        if not cand.exists():
            return cand
        i += 1

def _resolve_xlsm_path(path_like: str) -> Path:
    p = Path(path_like).expanduser().resolve()
    if p.suffix == "":
        p = p.with_suffix(".xlsm")
    if p.suffix.lower() != ".xlsm":
        raise SystemExit(f"Only .xlsm files are supported (got {p.suffix}).")
    return p

class Step:
    def __init__(self, name: str, path: Optional[Path], kind: str = "python_script"):
        self.name = name
        self.path = path
        self.kind = kind
    def exists(self) -> bool:
        return self.kind == "internal" or (self.path and self.path.exists())

def _discover_steps(project_root: Path) -> List[Step]:
    steps: List[Step] = []
    ctj = project_root / "convertToJson.py"
    steps.append(Step("convertToJson.py", ctj))
    perror = project_root / "Component" / "printerError" / "printerError.py"
    steps.append(Step("Component/printerError/printerError.py", perror))
    tf_dir = project_root / "Toner-finder"
    if tf_dir.exists():
        for p in sorted(tf_dir.glob("*.py")):
            if p.name.startswith("_"):
                continue
            steps.append(Step(f"Toner-finder/{p.name}", p))
    ttype = project_root / "Toner-type" / "tonerType.py"
    steps.append(Step("Toner-type/tonerType.py", ttype))
    steps.append(Step("Update Excel (internal)", None, kind="internal"))
    return [s for s in steps if s.exists()]

def _enumerate_steps(steps: List[Step]) -> List[Tuple[str, Step]]:
    out: List[Tuple[str, Step]] = []
    tcounter = 0
    for s in steps:
        if s.name == "convertToJson.py":
            out.append(("1", s))
        elif s.name == "Component/printerError/printerError.py":
            out.append(("2", s))
        elif s.name.startswith("Toner-finder/"):
            tcounter += 1
            out.append((f"3.{tcounter}", s))
        elif s.name == "Toner-type/tonerType.py":
            out.append(("4", s))
        elif s.kind == "internal":
            out.append(("5", s))
        else:
            out.append((str(len(out) + 1), s))
    return out

def _normalize(s: str) -> str:
    return s.lower().replace("\\", "/")

def _resolve_excludes(tokens: List[str], numbered: List[Tuple[str, Step]]) -> Set[int]:
    excludes: Set[int] = set()
    names = [_normalize(s.name) for _, s in numbered]
    num_map = {n: i for i, (n, _) in enumerate(numbered)}
    for t in tokens:
        t = str(t).strip()
        if not t:
            continue
        nt = _normalize(t)
        if nt in num_map:
            excludes.add(num_map[nt])
            continue
        if nt.endswith("."):
            for k, i in num_map.items():
                if k.startswith(nt):
                    excludes.add(i)
            continue
        if nt.isdigit() and nt in num_map:
            excludes.add(num_map[nt])
            continue
        prefix_matches = [i for i, n in enumerate(names) if n.split("/")[-1].startswith(nt) or n.startswith(nt)]
        if len(prefix_matches) == 1:
            excludes.add(prefix_matches[0])
            continue
        if len(prefix_matches) > 1:
            print(f"[exclude] '{t}' is ambiguous, matches: " + ", ".join(f"{numbered[i][0]}:{numbered[i][1].name}" for i in prefix_matches))
            continue
        sub_matches = [i for i, n in enumerate(names) if nt in n]
        if len(sub_matches) == 1:
            excludes.add(sub_matches[0])
        elif len(sub_matches) > 1:
            print(f"[exclude] '{t}' is ambiguous, matches: " + ", ".join(f"{numbered[i][0]}:{numbered[i][1].name}" for i in sub_matches))
        else:
            print(f"[exclude] No match for '{t}'.")
    return excludes

def _print_steps(numbered: List[Tuple[str, Step]]) -> None:
    print("Discovered steps:")
    for num, s in numbered:
        tag = " (internal)" if s.kind == "internal" else ""
        print(f" {num:>4} {s.name}{tag}")

def _run_steps(numbered: List[Tuple[str, Step]], excludes: Set[int], dry_run: bool) -> None:
    for i, (num, s) in enumerate(numbered):
        if i in excludes:
            print(f"[skip] {num}:{s.name}")
            continue
        print(f"[run ] {num}:{s.name}")
        if dry_run:
            continue
        if s.kind == "internal":
            continue
        try:
            subprocess.run([sys.executable, str(s.path)], check=True)
        except subprocess.CalledProcessError as e:
            print(f"[error] {num}:{s.name} exited with {e.returncode}")
        except Exception as e:
            print(f"[error] {num}:{s.name} failed: {e}")

def main(json_path: str, xlsm_path: str, update: bool = True, log: bool = True, pipeline: bool = False, exclude_tokens: Optional[List[str]] = None, list_steps: bool = False, dry_run: bool = False) -> int:
    project_root = _script_base_dir()
    if pipeline or list_steps:
        steps = _discover_steps(project_root)
        numbered = _enumerate_steps(steps)
        if list_steps:
            _print_steps(numbered)
            return 0
        excludes = _resolve_excludes(exclude_tokens or [], numbered)
        _print_steps(numbered)
        if excludes:
            print("Excluding:", ", ".join(numbered[i][0] for i in sorted(excludes)))
        _run_steps(numbered, excludes, dry_run=dry_run)
        if dry_run:
            print("[dry-run] Skipping Excel update because of --dry-run.")
            return 0
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    id_map = build_id_map(data)
    xlsm_path = str(_resolve_xlsm_path(xlsm_path))
    wb = openpyxl.load_workbook(filename=xlsm_path, data_only=False, keep_vba=True)
    total_updates = 0
    for ws in wb.worksheets:
        total_updates += update_sheet(ws, id_map)
    employees_path = _default_employees_json(project_root)
    if 'kwargs_employees_path' in globals() and kwargs_employees_path:
        employees_path = kwargs_employees_path
    emp_updates = 0
    try:
        if employees_path.exists():
            with open(employees_path, "r", encoding="utf-8") as ef:
                employees_json_obj = json.load(ef)
            employees_index = build_employees_index(employees_json_obj)
            print(f"[employees] index size: {len(employees_index)} from {employees_path}")
            if "Branches_Grouped" in wb.sheetnames:
                ws_bg = wb["Branches_Grouped"]
                emp_updates = update_branches_grouped(ws_bg, employees_index)
            else:
                print("[employees] sheet 'Branches_Grouped' not found")
        else:
            print(f"[employees] not found: {employees_path}")
    except Exception as e:
        print(f"[employees] error: {e}")
    total_updates += emp_updates
    print(f"[employees] updated rows: {emp_updates}")
    saved = []
    if update:
        wb.save(xlsm_path)
        saved.append(f"overwrite:{xlsm_path}")
    if log:
        logs_dir = _logs_dir()
        logs_dir.mkdir(parents=True, exist_ok=True)
        backup_name = _safe_timestamp_filename(".xlsm")
        backup_path = _unique_path(logs_dir, backup_name)
        wb.save(str(backup_path))
        saved.append(f"log:{backup_path}")
    if not update and not log:
        p = Path(xlsm_path)
        out_path = str(p.with_name(p.stem + "_updated" + p.suffix))
        wb.save(out_path)
        saved.append(f"copy:{out_path}")
    print(f"Updated rows: {total_updates}")
    for s in saved:
        print(f"Saved: {s}")
    return 0

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(add_help=True, description="Printer JSON → XLSM updater with optional troubleshooting runner.")
    parser.add_argument("json_path")
    parser.add_argument("xlsm_path")
    parser.add_argument("-u","--update", default="true")
    parser.add_argument("-l","--log", default="true")
    parser.add_argument("-p","--pipeline", action="store_true")
    parser.add_argument("-e","--exclude", nargs="*", default=[])
    parser.add_argument("--list-steps", action="store_true")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--employees-json", default="", help="Optional path to employeesData.json (defaults to data/employeesData.json)")
    args = parser.parse_args()
    def _to_bool(x):
        s = str(x).strip().lower()
        if s in ("1","true","t","yes","y","on"):
            return True
        if s in ("0","false","f","no","n","off"):
            return False
        return True
    up = _to_bool(args.update)
    lg = _to_bool(args.log)
    kwargs_employees_path = None
    if args.employees_json:
        try:
            ep = Path(args.employees_json).expanduser().resolve()
            kwargs_employees_path = ep
        except Exception:
            kwargs_employees_path = None
    sys.exit(main(
        args.json_path,
        args.xlsm_path,
        update=up,
        log=lg,
        pipeline=args.pipeline,
        exclude_tokens=args.exclude,
        list_steps=args.list_steps,
        dry_run=args.dry_run,
    ))
